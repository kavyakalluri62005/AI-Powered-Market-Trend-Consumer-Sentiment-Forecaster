# ============================================================
# WEEK 6 - Full Integration
# AI-Powered Market Trend & Consumer Sentiment Forecaster
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import warnings
warnings.filterwarnings("ignore")

# ── PAGE CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="Consumer Sentiment Forecaster",
    page_icon="🧠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CUSTOM CSS ───────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        font-weight: bold;
        text-align: center;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        padding: 0.5rem;
    }
    .week-badge {
        background: linear-gradient(135deg, #667eea, #764ba2);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: bold;
    }
    .alert-high   { background:#FADBD8; padding:10px; border-radius:8px; border-left:4px solid #E74C3C; margin:5px 0; }
    .alert-medium { background:#FDEBD0; padding:10px; border-radius:8px; border-left:4px solid #F39C12; margin:5px 0; }
    .alert-low    { background:#D5F5E3; padding:10px; border-radius:8px; border-left:4px solid #27AE60; margin:5px 0; }
</style>
""", unsafe_allow_html=True)

# ── HELPER FUNCTIONS ─────────────────────────────────────────
def to_ascii(text, limit=None):
    result = str(text).encode("ascii", "ignore").decode("ascii").strip()
    return result[:limit] if limit else result

@st.cache_data
def load_all_data():
    df        = pd.read_csv("sentiment_results.csv")
    topics_df = pd.read_csv("topics_summary.csv")
    rag_df    = pd.read_csv("rag_insights.csv")
    alerts_df = pd.read_csv("alerts_log.csv")

    if "date" in df.columns:
        df["date"]  = pd.to_datetime(df["date"], errors="coerce")
        df["year"]  = df["date"].dt.year
        df["month"] = df["date"].dt.month

    return df, topics_df, rag_df, alerts_df

def detect_alerts(df):
    alerts    = []
    total     = len(df)
    neg_count = len(df[df["ai_sentiment"] == "Negative"])
    neg_pct   = neg_count / total
    avg_score = df["sentiment_score"].mean()

    if neg_pct > 0.40:
        alerts.append({"type": "HIGH NEGATIVE SENTIMENT",
                       "message": f"Negative sentiment at {neg_pct:.1%} exceeds 40% threshold",
                       "severity": "HIGH"})
    if avg_score < 0.0:
        alerts.append({"type": "NEGATIVE SENTIMENT SCORE",
                       "message": f"Average score dropped to {avg_score:.3f}",
                       "severity": "MEDIUM"})
    if "year" in df.columns:
        yearly = df.groupby("year")["sentiment_score"].mean()
        for i in range(1, len(yearly)):
            drop = yearly.iloc[i-1] - yearly.iloc[i]
            if drop > 0.2:
                alerts.append({"type": f"SENTIMENT DROP {yearly.index[i-1]} to {yearly.index[i]}",
                               "message": f"Dropped by {drop:.3f}",
                               "severity": "MEDIUM"})
    if not alerts:
        alerts.append({"type": "ALL CLEAR",
                       "message": f"Sentiment healthy at {neg_pct:.1%} negative",
                       "severity": "LOW"})
    return alerts

def generate_excel(df, topics_df, rag_df, alerts):
    wb           = openpyxl.Workbook()
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2C3E50")
    pos_fill     = PatternFill("solid", fgColor="D5F5E3")
    neg_fill     = PatternFill("solid", fgColor="FADBD8")
    alert_fill   = PatternFill("solid", fgColor="FDEBD0")

    total     = len(df)
    neg_count = len(df[df["ai_sentiment"] == "Negative"])
    neg_pct   = neg_count / total
    avg_score = df["sentiment_score"].mean()

    # Sheet 1: Summary
    ws1       = wb.active
    ws1.title = "Executive Summary"
    ws1["A1"] = "AI-Powered Consumer Sentiment Forecaster - Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1.merge_cells("A1:D1")
    ws1.merge_cells("A2:D2")
    ws1["A4"] = "METRIC"
    ws1["B4"] = "VALUE"
    for col in [ws1["A4"], ws1["B4"]]:
        col.font = header_font
        col.fill = header_fill

    kpis = [
        ("Total Reviews", f"{total:,}"),
        ("Positive Reviews", f"{len(df[df['ai_sentiment']=='Positive']):,} ({1-neg_pct:.1%})"),
        ("Negative Reviews", f"{neg_count:,} ({neg_pct:.1%})"),
        ("Avg Sentiment Score", f"{avg_score:.3f}"),
        ("Avg Star Rating", f"{df['rating'].mean():.2f}/5.0"),
        ("Alerts Generated", f"{len(alerts)}"),
    ]
    for i, (m, v) in enumerate(kpis, 5):
        ws1.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=v)
        fill = pos_fill if i % 2 == 0 else PatternFill("solid", fgColor="F8F9FA")
        for c in [1, 2]:
            ws1.cell(row=i, column=c).fill = fill
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    # Sheet 2: Data
    ws2 = wb.create_sheet("Sentiment Data")
    cols = [c for c in ["review_text", "ai_sentiment", "ai_confidence", "sentiment_score", "rating"] if c in df.columns]
    for ci, cn in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=ci, value=cn.upper())
        cell.font = header_font
        cell.fill = header_fill
    for ri, row in enumerate(df[cols].head(500).itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=str(val)[:200])
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 25

    # Sheet 3: Alerts
    ws3 = wb.create_sheet("Alerts")
    for ci, h in enumerate(["Type", "Message", "Severity"], 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for ri, alert in enumerate(alerts, 2):
        ws3.cell(row=ri, column=1, value=alert["type"])
        ws3.cell(row=ri, column=2, value=alert["message"])
        ws3.cell(row=ri, column=3, value=alert["severity"])
        fill = neg_fill if alert["severity"] == "HIGH" else alert_fill
        for c in [1, 2, 3]:
            ws3.cell(row=ri, column=c).fill = fill
    for col in ws3.columns:
        ws3.column_dimensions[get_column_letter(col[0].column)].width = 35

    # Sheet 4: Topics
    ws4 = wb.create_sheet("Topics")
    for ci, h in enumerate(["Topic ID", "Keywords", "Count"], 1):
        cell = ws4.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for ri, row in enumerate(topics_df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws4.cell(row=ri, column=ci, value=str(val))
    for col in ws4.columns:
        ws4.column_dimensions[get_column_letter(col[0].column)].width = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf(df, topics_df, rag_df, alerts):
    total     = len(df)
    neg_count = len(df[df["ai_sentiment"] == "Negative"])
    neg_pct   = neg_count / total
    avg_score = df["sentiment_score"].mean()

    class SentimentPDF(FPDF):
        def header(self):
            self.set_fill_color(44, 62, 80)
            self.rect(0, 0, 210, 20, "F")
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(255, 255, 255)
            self.set_y(5)
            self.cell(0, 10, "AI-Powered Consumer Sentiment Forecaster", align="C", ln=True)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f"Page {self.page_no()} | {datetime.now().strftime('%Y-%m-%d')}", align="C")

        def section_title(self, title):
            self.ln(5)
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(41, 128, 185)
            self.cell(0, 10, to_ascii(title), ln=True)
            self.ln(2)

        def kpi_row(self, label, value, r, g, b):
            self.set_fill_color(r, g, b)
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(44, 62, 80)
            self.cell(95, 9, f"  {to_ascii(label)}", fill=True)
            self.set_font("Helvetica", "", 10)
            self.cell(85, 9, f"  {to_ascii(value)}", fill=True, ln=True)
            self.ln(1)

    pdf = SentimentPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(15, 25, 15)

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(44, 62, 80)
    pdf.cell(0, 12, "Executive Report", align="C", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(127, 140, 141)
    pdf.cell(0, 8, f"Generated on {datetime.now().strftime('%B %d, %Y')}", align="C", ln=True)
    pdf.ln(6)

    pdf.section_title("Key Performance Indicators")
    pdf.kpi_row("Total Reviews Analyzed", f"{total:,}", 236, 240, 241)
    pdf.kpi_row("Positive Reviews", f"{len(df[df['ai_sentiment']=='Positive']):,} ({1-neg_pct:.1%})", 213, 245, 227)
    pdf.kpi_row("Negative Reviews", f"{neg_count:,} ({neg_pct:.1%})", 250, 219, 216)
    pdf.kpi_row("Avg Sentiment Score", f"{avg_score:.3f}", 236, 240, 241)
    pdf.kpi_row("Avg Star Rating", f"{df['rating'].mean():.2f}/5.0", 236, 240, 241)
    pdf.kpi_row("Total Alerts", f"{len(alerts)}", 236, 240, 241)

    pdf.section_title("Alerts and Recommendations")
    for alert in alerts:
        if alert["severity"] == "HIGH":
            pdf.set_fill_color(250, 219, 216)
        elif alert["severity"] == "MEDIUM":
            pdf.set_fill_color(253, 235, 208)
        else:
            pdf.set_fill_color(213, 245, 227)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(44, 62, 80)
        pdf.cell(0, 8, f"  [{to_ascii(alert['severity'])}] {to_ascii(alert['type'], 80)}", fill=True, ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(60, 60, 60)
        pdf.cell(0, 7, f"  {to_ascii(alert['message'], 120)}", fill=True, ln=True)
        pdf.ln(2)

    pdf.section_title("Top Discovered Topics")
    for _, row in topics_df.head(5).iterrows():
        pdf.set_fill_color(236, 240, 241)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(44, 62, 80)
        pdf.cell(0, 8, f"  Topic {int(row['topic_id'])}: {to_ascii(row['top_words'], 70)}", fill=True, ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 7, f"  Reviews: {int(row['count'])}", ln=True)
        pdf.ln(1)

    if not rag_df.empty:
        pdf.section_title("AI Consumer Insights (RAG Pipeline)")
        for _, row in rag_df.head(3).iterrows():
            pdf.set_fill_color(214, 234, 248)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(0, 8, f"  Q: {to_ascii(row['question'], 100)}", fill=True, ln=True)
            pdf.set_fill_color(245, 250, 254)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 7, f"  A: {to_ascii(row['answer'], 200)}", fill=True, ln=True)
            pdf.ln(3)

    return bytes(pdf.output())


# ── LOAD DATA ────────────────────────────────────────────────
df, topics_df, rag_df, alerts_df = load_all_data()
alerts = detect_alerts(df)

# ── SIDEBAR ──────────────────────────────────────────────────
st.sidebar.image("https://img.icons8.com/fluency/96/combo-chart.png", width=70)
st.sidebar.title("Sentiment Forecaster")
st.sidebar.markdown("---")

st.sidebar.subheader("Filters")
sentiment_filter = st.sidebar.multiselect(
    "Sentiment", options=["Positive", "Negative"],
    default=["Positive", "Negative"]
)

if "year" in df.columns:
    years      = sorted(df["year"].dropna().unique().astype(int).tolist())
    year_range = st.sidebar.select_slider("Year Range", options=years, value=(min(years), max(years)))
else:
    year_range = None

rating_filter = st.sidebar.multiselect(
    "Star Rating", options=[1, 2, 3, 4, 5], default=[1, 2, 3, 4, 5]
)

st.sidebar.markdown("---")
st.sidebar.subheader("Download Reports")

excel_data = generate_excel(df, topics_df, rag_df, alerts)
st.sidebar.download_button(
    label="📥 Download Excel Report",
    data=excel_data,
    file_name=f"sentiment_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

pdf_data = generate_pdf(df, topics_df, rag_df, alerts)
st.sidebar.download_button(
    label="📥 Download PDF Report",
    data=pdf_data,
    file_name=f"sentiment_report_{datetime.now().strftime('%Y%m%d')}.pdf",
    mime="application/pdf",
    use_container_width=True
)

st.sidebar.markdown("---")
st.sidebar.markdown("**Built With:**")
st.sidebar.markdown("HuggingFace • LangChain • Groq • FAISS • Streamlit")

# Apply filters
filtered_df = df[df["ai_sentiment"].isin(sentiment_filter)]
if "rating" in df.columns:
    filtered_df = filtered_df[filtered_df["rating"].isin(rating_filter)]
if year_range and "year" in df.columns:
    filtered_df = filtered_df[
        (filtered_df["year"] >= year_range[0]) &
        (filtered_df["year"] <= year_range[1])
    ]

# ── HEADER ───────────────────────────────────────────────────
st.markdown('<p class="main-header">🧠 AI-Powered Market Trend & Consumer Sentiment Forecaster</p>', unsafe_allow_html=True)
st.markdown("---")

# ── ALERT BANNER ─────────────────────────────────────────────
high_alerts = [a for a in alerts if a["severity"] == "HIGH"]
if high_alerts:
    for a in high_alerts:
        st.error(f"🚨 **{a['type']}**: {a['message']}")
elif any(a["severity"] == "MEDIUM" for a in alerts):
    for a in [x for x in alerts if x["severity"] == "MEDIUM"]:
        st.warning(f"⚠️ **{a['type']}**: {a['message']}")
else:
    st.success("✅ All Clear — Sentiment is healthy!")

# ── TABS ─────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📈 Overview",
    "💬 Sentiment Analysis",
    "🏷️ Topic Modeling",
    "🤖 AI Insights (RAG)",
    "🚨 Alerts & Reports"
])

# ════════════════════════════════════════════════════════════
# TAB 1: OVERVIEW
# ════════════════════════════════════════════════════════════
with tab1:
    st.subheader("Dataset Overview")
    col1, col2, col3, col4 = st.columns(4)
    total     = len(filtered_df)
    pos_count = len(filtered_df[filtered_df["ai_sentiment"] == "Positive"])
    neg_count = len(filtered_df[filtered_df["ai_sentiment"] == "Negative"])
    avg_score = filtered_df["sentiment_score"].mean()

    col1.metric("Total Reviews",     f"{total:,}")
    col2.metric("Positive Reviews",  f"{pos_count:,}", f"{pos_count/total:.1%}")
    col3.metric("Negative Reviews",  f"{neg_count:,}", f"-{neg_count/total:.1%}")
    col4.metric("Avg Sentiment Score", f"{avg_score:.3f}")

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Rating Distribution")
        rating_counts = filtered_df["rating"].value_counts().sort_index()
        fig = px.bar(x=rating_counts.index, y=rating_counts.values,
                     color=rating_counts.index, color_continuous_scale="RdYlGn",
                     labels={"x": "Star Rating", "y": "Count"})
        fig.update_layout(showlegend=False, height=320)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Sentiment Distribution")
        sc = filtered_df["ai_sentiment"].value_counts()
        fig = px.pie(values=sc.values, names=sc.index,
                     color=sc.index,
                     color_discrete_map={"Positive": "#2ecc71", "Negative": "#e74c3c"},
                     hole=0.4)
        fig.update_layout(height=320)
        st.plotly_chart(fig, use_container_width=True)

    if "year" in filtered_df.columns:
        st.subheader("Reviews Over Time")
        yearly = filtered_df.groupby(["year", "ai_sentiment"]).size().reset_index(name="count")
        fig = px.line(yearly, x="year", y="count", color="ai_sentiment",
                      color_discrete_map={"Positive": "#2ecc71", "Negative": "#e74c3c"},
                      markers=True)
        fig.update_layout(height=320)
        st.plotly_chart(fig, use_container_width=True)

# ════════════════════════════════════════════════════════════
# TAB 2: SENTIMENT ANALYSIS
# ════════════════════════════════════════════════════════════
with tab2:
    st.subheader("Deep Sentiment Analysis")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("AI Confidence Scores")
        fig = px.histogram(filtered_df, x="ai_confidence", color="ai_sentiment",
                           color_discrete_map={"Positive": "#2ecc71", "Negative": "#e74c3c"},
                           nbins=30)
        fig.update_layout(height=320)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Sentiment Score by Rating")
        fig = px.box(filtered_df, x="rating", y="sentiment_score", color="ai_sentiment",
                     color_discrete_map={"Positive": "#2ecc71", "Negative": "#e74c3c"})
        fig.update_layout(height=320)
        st.plotly_chart(fig, use_container_width=True)

    if "year" in filtered_df.columns:
        st.subheader("Sentiment Score Trend Over Time")
        trend = filtered_df.groupby("year")["sentiment_score"].mean().reset_index()
        fig = px.area(trend, x="year", y="sentiment_score",
                      color_discrete_sequence=["#667eea"])
        fig.add_hline(y=0, line_dash="dash", line_color="red", annotation_text="Neutral")
        fig.update_layout(height=320)
        st.plotly_chart(fig, use_container_width=True)

    st.subheader("Word Cloud")
    wc_col1, wc_col2 = st.columns([1, 3])
    with wc_col1:
        wc_sentiment = st.radio("Select", ["Positive", "Negative"])
    wc_texts = " ".join(
        filtered_df[filtered_df["ai_sentiment"] == wc_sentiment]["review_text"]
        .dropna().astype(str).tolist()
    )
    if wc_texts:
        wc = WordCloud(width=800, height=300, background_color="white",
                       colormap="Greens" if wc_sentiment == "Positive" else "Reds",
                       max_words=100).generate(wc_texts[:50000])
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.imshow(wc, interpolation="bilinear")
        ax.axis("off")
        st.pyplot(fig)

# ════════════════════════════════════════════════════════════
# TAB 3: TOPIC MODELING
# ════════════════════════════════════════════════════════════
with tab3:
    st.subheader("Discovered Topics")
    if not topics_df.empty:
        col1, col2 = st.columns(2)
        with col1:
            fig = px.bar(topics_df, x="count",
                         y=topics_df["topic_id"].astype(str),
                         orientation="h", color="count",
                         color_continuous_scale="Purples")
            fig.update_layout(height=400, showlegend=False)
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.subheader("Topic Keywords")
            for _, row in topics_df.iterrows():
                with st.expander(f"Topic {int(row['topic_id'])} — {str(row['top_words'])[:40]}"):
                    st.write(f"**Keywords:** {row['top_words']}")
                    st.write(f"**Reviews:** {int(row['count'])}")

# ════════════════════════════════════════════════════════════
# TAB 4: RAG AI INSIGHTS
# ════════════════════════════════════════════════════════════
with tab4:
    st.subheader("AI Consumer Insights (RAG Pipeline)")
    if not rag_df.empty:
        st.success(f"✅ {len(rag_df)} AI insights generated")
        for _, row in rag_df.iterrows():
            with st.expander(f"❓ {row['question']}"):
                st.markdown(f"**🤖 AI Answer:**\n\n{row['answer']}")
    else:
        st.info("No RAG insights found.")

    st.markdown("---")
    st.subheader("Platform Architecture")
    col1, col2, col3, col4 = st.columns(4)
    col1.success("**Week 1**\nData Pipeline\n568K reviews")
    col2.success("**Week 2**\nAI Sentiment\n5K analyzed")
    col3.success("**Week 3**\nRAG Pipeline\nContextual AI")
    col4.success("**Week 4**\nDashboard\nLive charts")

# ════════════════════════════════════════════════════════════
# TAB 5: ALERTS & REPORTS
# ════════════════════════════════════════════════════════════
with tab5:
    st.subheader("Alerts & Reports")

    st.subheader("Active Alerts")
    for alert in alerts:
        severity = alert["severity"]
        if severity == "HIGH":
            st.markdown(f'<div class="alert-high">🚨 <b>[HIGH]</b> {alert["type"]}<br>{alert["message"]}</div>',
                        unsafe_allow_html=True)
        elif severity == "MEDIUM":
            st.markdown(f'<div class="alert-medium">⚠️ <b>[MEDIUM]</b> {alert["type"]}<br>{alert["message"]}</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="alert-low">✅ <b>[LOW]</b> {alert["type"]}<br>{alert["message"]}</div>',
                        unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("Download Reports")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 📊 Excel Report")
        st.write("Contains 4 sheets: Summary, Sentiment Data, Alerts, Topics")
        excel_data2 = generate_excel(df, topics_df, rag_df, alerts)
        st.download_button(
            label="📥 Download Excel",
            data=excel_data2,
            file_name=f"sentiment_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col2:
        st.markdown("### 📄 PDF Report")
        st.write("Professional executive report with KPIs, alerts and insights")
        pdf_data2 = generate_pdf(df, topics_df, rag_df, alerts)
        st.download_button(
            label="📥 Download PDF",
            data=pdf_data2,
            file_name=f"sentiment_report_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    st.markdown("---")
    st.subheader("Raw Alerts Log")
    st.dataframe(pd.DataFrame(alerts), use_container_width=True)
